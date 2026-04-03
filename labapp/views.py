from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from django.urls import reverse
from django.db import DatabaseError, IntegrityError, transaction
from .models import College, CollegeAdmin, Lab, Professor, Student, Division, Experiment, Submission, Evaluation, Attendance, VivaSession, ExcelUpload
import json
import openpyxl
from io import BytesIO
from datetime import datetime
import logging
import mimetypes
import os
import re

logger = logging.getLogger(__name__)

DOCUMENT_EXTENSIONS = (
    '.pdf', '.doc', '.docx', '.ppt', '.pptx',
    '.xls', '.xlsx', '.txt', '.csv', '.zip', '.rar'
)


def _safe_worksheet_title(title, fallback='Sheet1'):
    """
    Excel sheet titles cannot contain []:*?/\\ and are limited to 31 chars.
    """
    cleaned = re.sub(r'[\[\]\*\?/:\\]', ' ', str(title or '').strip())
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return (cleaned or fallback)[:31]


def _safe_filename_part(value, fallback='file'):
    """
    Keep download filenames cross-platform safe and predictable.
    """
    cleaned = re.sub(r'[^A-Za-z0-9._-]+', '_', str(value or '').strip())
    cleaned = cleaned.strip('._')
    return cleaned or fallback


def _resource_url_for_download(file_field):
    """
    Build a downloadable URL for uploaded resources.
    For Cloudinary, document files should use /raw/upload/ instead of /image/upload/.
    """
    if not file_field:
        return None

    try:
        url = file_field.url
    except Exception:
        return None

    file_name = (getattr(file_field, 'name', '') or '').lower()
    is_document = file_name.endswith(DOCUMENT_EXTENSIONS)
    if is_document and '/image/upload/' in url:
        return url.replace('/image/upload/', '/raw/upload/')
    return url


def _resource_exists(file_field):
    """True when a file reference has a backing object in its storage."""
    try:
        if not file_field:
            return False
        storage = getattr(file_field, 'storage', None)
        name = getattr(file_field, 'name', '')
        if not storage or not name:
            return False
        try:
            exists = storage.exists(name)
        except Exception:
            exists = False

        if exists:
            return True

        # Some storage backends may not accurately report exists(). Fall back to URL.
        try:
            return bool(getattr(file_field, 'url', None))
        except Exception:
            return False
    except Exception:
        return False


def _safe_field_url(file_field):
    """Return a usable URL for a FileField/ImageField if possible."""
    try:
        if not file_field:
            return ''
        return file_field.url
    except Exception:
        return ''


def _build_lab_resource_url(lab_id, resource_type, download=False):
    """Return app endpoint URL for serving lab resources."""
    url = reverse('student_download_lab_resource', args=[lab_id, resource_type])
    if download:
        return f"{url}?download=1"
    return url


def _file_response_from_field(file_field, download=False):
    """
    Build a streaming response from storage (works for local and Cloudinary backends).
    Falls back to redirect URL only if direct open fails.
    """
    if not file_field:
        return HttpResponse("Requested file is not available.", status=404)

    filename = os.path.basename(getattr(file_field, 'name', '') or 'download')
    content_type, _ = mimetypes.guess_type(filename)
    disposition = 'attachment' if download else 'inline'

    try:
        file_field.open('rb')
        response = FileResponse(file_field, content_type=content_type or 'application/octet-stream')
        response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
        return response
    except Exception:
        logger.exception("Failed to stream lab resource from storage: %s", filename)
        # Fallback for storages that prefer URL-based access.
        resource_url = _resource_url_for_download(file_field)
        if resource_url:
            return redirect(resource_url)
        return HttpResponse("Unable to access requested file. Please re-upload from professor panel.", status=404)


def _submission_access_allowed(request, submission):
    """Return True when the requester can access a submission's media."""
    is_owner_student = Student.objects.filter(user=request.user, id=submission.student_id).exists()

    sub_lab = submission.lab if submission.lab else (
        submission.experiment.lab if submission.experiment else None
    )
    is_owner_professor = False
    if sub_lab is not None:
        is_owner_professor = Professor.objects.filter(
            user=request.user,
            id=sub_lab.professor_id
        ).exists()

    return is_owner_student or is_owner_professor

def home(request):
    return render(request, "base/home.html")

@login_required
def student_dashboard(request):
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        return redirect('college_auth') # Or wherever appropriate

    # 1. Fetch Explicitly Assigned Labs
    labs = student.assigned_labs.all()

    # 2. Fetch Attendance Data
    attendance_data = []
    for lab in labs:
        attendances = Attendance.objects.filter(student=student, lab=lab).order_by('date')
        
        dates = [a.date.strftime('%d %b') for a in attendances]
        status_map = {a.date.strftime('%d %b'): ('present' if a.present else 'absent') for a in attendances}
        present_count = attendances.filter(present=True).count()
        total_count = attendances.count()
        percentage = round((present_count / total_count) * 100) if total_count > 0 else 0
        
        attendance_data.append({
            'lab_id': lab.id,
            'lab_name': lab.name,
            'dates': dates,
            'status_map': status_map,
            'percentage': percentage
        })

    # 3. Fetch Marks Data
    marks_data = []
    for lab in labs:
        submissions = Submission.objects.filter(student=student, experiment__lab=lab).select_related('evaluation', 'experiment')
        records = []
        lab_total = 0
        exp_count = 0

        for sub in submissions:
            if hasattr(sub, 'evaluation'):
                viva = sub.evaluation.viva_marks
                exp_marks = sub.evaluation.experiment_marks
                writeup = sub.evaluation.writeup_marks
                total = viva + exp_marks + writeup
                lab_total += total
                exp_count += 1
                records.append({
                    'exp_number': sub.experiment.number,
                    'viva': viva,
                    'exp_marks': exp_marks,
                    'writeup': writeup,
                    'total': total
                })

        average = round(lab_total / exp_count, 2) if exp_count > 0 else 0
        max_possible = exp_count * 15
        percentage = round((lab_total / max_possible) * 100, 2) if max_possible > 0 else 0

        marks_data.append({
            'lab_id': lab.id,
            'lab_name': lab.name,
            'records': records,
            'grand_total': lab_total,
            'average': average,
            'percentage': percentage
        })

    # 4. Fetch Submission History (only new-style submissions with experiment_name)
    history_data = []
    for lab in labs:
        submissions = Submission.objects.filter(
            student=student,
            lab=lab
        ).exclude(experiment_name='').order_by('-submitted_at')
        history_data.append({
            'lab_id': lab.id,
            'lab_name': lab.name,
            'submissions': submissions
        })

    # 5. Handle Upload Experiment (POST)
    if request.method == "POST":
        lab_id = request.POST.get('lab_id')
        experiment_name = request.POST.get('experiment_name', '').strip()
        code_screenshot = request.FILES.get('codeScreenshot')
        output_screenshot = request.FILES.get('outputScreenshot')

        if experiment_name and code_screenshot and output_screenshot and lab_id:
            try:
                lab = Lab.objects.get(id=lab_id)
                Submission.objects.create(
                    student=student,
                    lab=lab,
                    experiment_name=experiment_name,
                    code_screenshot=code_screenshot,
                    output_screenshot=output_screenshot,
                    status='pending'
                )
                messages.success(request, "Experiment submitted successfully!")
            except Lab.DoesNotExist:
                messages.error(request, "Selected lab not found.")
        else:
            messages.error(request, "Please fill in all fields.")
        return redirect('student_dashboard')

    # Check for active viva sessions
    active_session = VivaSession.objects.filter(student=student, is_active=True).first()

    context = {
        'student': student,
        'labs': labs,
        'attendance_data': attendance_data,
        'marks_data': marks_data,
        'history_data': history_data,
        'active_session': active_session,
    }

    return render(request, "student/student_dashboard.html", context)

@login_required
def update_profile(request):
    if request.method == "POST":
        try:
            student = Student.objects.get(user=request.user)
            user = request.user
            
            if 'profile_pic' in request.FILES:
                student.profile_pic = request.FILES['profile_pic']
                student.save()
            
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if new_password and confirm_password:
                if new_password == confirm_password:
                    user.set_password(new_password)
                    user.save()
                    update_session_auth_hash(request, user)  # Keeps user logged in after password change
                    messages.success(request, "Password updated successfully!")
                else:
                    messages.error(request, "Passwords do not match!")
            else:
                messages.success(request, "Profile updated successfully!")
                
        except Student.DoesNotExist:
            messages.error(request, "Student profile not found.")
            
    return redirect('student_dashboard')

@login_required
def student_logout(request):
    logout(request)
    return redirect('student_login')

@login_required
def export_marks_excel(request, lab_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    try:
        student = Student.objects.get(user=request.user)
        lab = Lab.objects.get(id=lab_id)
    except (Student.DoesNotExist, Lab.DoesNotExist):
        return HttpResponse("Student or Lab not found.", status=404)

    # Need data for up to 10 experiments ideally or just those submitted
    submissions = Submission.objects.filter(student=student, experiment__lab=lab).select_related('evaluation', 'experiment').order_by('experiment__number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Marks - {lab.name}"

    # Header Row
    ws.merge_cells('A1:B1')
    ws['A1'] = "Student Details"
    ws['A1'].font = Font(bold=True)
    ws['C1'] = f"Name: {student.user.username}"
    ws['D1'] = f"PRN: {student.prn}"

    headers = ["Exp No", "Experiment (5 M)", "Write-up (5 M)", "Viva (5 M)", "Total (15 M)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 4
    total_obtained = 0
    total_possible = 0
    
    for sub in submissions:
        if hasattr(sub, 'evaluation'):
            viva = float(sub.evaluation.viva_marks or 0)
            exp_marks = float(sub.evaluation.experiment_marks or 0)
            writeup = float(sub.evaluation.writeup_marks or 0)
            total = viva + exp_marks + writeup
            
            ws.cell(row=row_num, column=1, value=sub.experiment.number).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=2, value=exp_marks).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3, value=writeup).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4, value=viva).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=5, value=total).alignment = Alignment(horizontal='center')
            
            total_obtained += total
            total_possible += 15 # Assuming max 15 per experiment
            row_num += 1

    # Add summary rows
    row_num += 1
    avg = round(total_obtained / max(1, (total_possible/15)), 2)
    percent = round((total_obtained / max(1, total_possible)) * 100, 2)
    
    ws.cell(row=row_num, column=4, value="Grand Total:").font = Font(bold=True)
    ws.cell(row=row_num, column=5, value=total_obtained)
    
    ws.cell(row=row_num+1, column=4, value="Average:").font = Font(bold=True)
    ws.cell(row=row_num+1, column=5, value=avg)

    ws.cell(row=row_num+2, column=4, value="Percentage:").font = Font(bold=True)
    ws.cell(row=row_num+2, column=5, value=f"{percent}%")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="marks_{student.prn}_{lab.name}.xlsx"'
    wb.save(response)
    
    return response

def student_login(request):
    active_tab = "login"

    if request.method == "POST":
        # LOGIN
        if "student_login_submit" in request.POST:
            identifier = request.POST.get("username")  # Can be username or PRN from the new form
            password = request.POST.get("password")
            
            # 1. Try direct authentication with the provided identifier
            user = authenticate(request, username=identifier, password=password)

            # 2. If it fails, check if the identifier is a PRN
            if user is None:
                try:
                    student_by_prn = Student.objects.get(prn=identifier)
                    # If PRN exists, try authenticating with the linked user's actual username
                    user = authenticate(request, username=student_by_prn.user.username, password=password)
                except Student.DoesNotExist:
                    pass

            if user is not None and hasattr(user, "student"):
                login(request, user)
                return redirect("student_dashboard")
            else:
                # Check if user exists but password is wrong OR if user doesn't exist at all
                messages.error(request, "Invalid username/PRN or password")

        # REGISTER
        elif "student_register_submit" in request.POST:
            active_tab = "register"

            username = request.POST.get("username")
            prn = request.POST.get("prn")
            password1 = request.POST.get("password1")
            password2 = request.POST.get("password2")
            division_name = request.POST.get("division")

            if password1 != password2:
                messages.error(request, "Passwords do not match")
                return render(request, "student/student_login.html", {"active_tab": active_tab})

            # Check if student PRN was uploaded by professor
            try:
                student = Student.objects.get(prn=prn)
            except Student.DoesNotExist:
                messages.error(request, "PRN does not exist. Please contact your professor or ensure it was uploaded in the Excel sheet.")
                return render(request, "student/student_login.html", {"active_tab": active_tab})
            
            # Check if this PRN already has a password set (is already registered fully)
            # The excel upload script creates a basic user with username=prn.lower() but no usable password.
            if student.user.has_usable_password():
                messages.error(request, "This PRN is already registered. Please go to the Sign In page.")
                return redirect("student_login")
            
            # If username changed, we should probably check if the new chosen username is taken by someone else
            if student.user.username != username and User.objects.filter(username=username).exists():
                messages.error(request, "Username already taken. Please choose a different ERP username or use your PRN.")
                return render(request, "student/student_login.html", {"active_tab": active_tab})

            try:
                # Update the existing user correctly
                user = student.user
                user.username = username
                user.set_password(password1)
                user.save()

                # Sync the division as well during registration if they picked a new one
                if division_name:
                    division_obj, _ = Division.objects.get_or_create(name=division_name, college=student.division.college)
                    student.division = division_obj
                    student.save()

                messages.success(request, "Student account registered successfully. Please sign in.")
                active_tab = "login"
            except Exception as e:
                messages.error(request, f"Error completing registration: {str(e)}")

    return render(request, "student/student_login.html", {"active_tab": active_tab})

def college_auth(request):

    if request.method == "POST":

        # LOGIN
        if "login_submit" in request.POST:

            username = request.POST.get("username", "").strip()
            password = request.POST.get("password")

            try:
                user = authenticate(request, username=username, password=password)
            except DatabaseError:
                messages.error(request, "Database is not ready. Run migrations and verify DATABASE_URL.")
                return redirect("college_auth")

            # Username login is case-sensitive by default in Django.
            # Try a case-insensitive username lookup for friendlier login behavior.
            if user is None and username:
                try:
                    matched_user = User.objects.get(username__iexact=username)
                    user = authenticate(request, username=matched_user.username, password=password)
                except User.DoesNotExist:
                    pass
                except DatabaseError:
                    messages.error(request, "Database is not ready. Run migrations and verify DATABASE_URL.")
                    return redirect("college_auth")

            if user is not None:
                try:
                    _ = user.collegeadmin
                    login(request, user)
                    return redirect("college_dashboard")
                except ObjectDoesNotExist:
                    messages.error(request, "This account is not registered as a college admin.")
                    return redirect("college_auth")

            messages.error(request, "Invalid username or password")

        # REGISTER
        elif "register_submit" in request.POST:

            college_name = request.POST.get("college_name", "").strip()
            college_email = request.POST.get("college_email", "").strip()
            username = request.POST.get("username", "").strip()
            password = request.POST.get("password")
            confirm_password = request.POST.get("confirm_password")

            if not college_name or not college_email or not username:
                messages.error(request, "College name, college email, and username are required.")
                return redirect("college_auth")

            if password != confirm_password:
                messages.error(request, "Passwords do not match")
                return redirect("college_auth")

            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=username,
                        password=password
                    )

                    college = College.objects.create(
                        name=college_name,
                        email=college_email
                    )

                    CollegeAdmin.objects.create(
                        user=user,
                        college=college
                    )
            except IntegrityError as exc:
                error_text = str(exc).lower()
                if "username" in error_text:
                    messages.error(request, "Username already exists. Please choose another username.")
                elif "email" in error_text:
                    messages.error(request, "College email already exists. Please use another email.")
                else:
                    messages.error(request, "Could not create account due to duplicate data.")
                return redirect("college_auth")
            except DatabaseError:
                messages.error(request, "Database is not ready. Run migrations and verify DATABASE_URL.")
                return redirect("college_auth")

            messages.success(request, "Account created successfully. Please login.")
            return redirect("college_auth")

    return render(request, "college/auth.html")

@login_required
def college_dashboard(request):

    try:
        college_admin = request.user.collegeadmin
    except ObjectDoesNotExist:
        messages.error(request, "Your account is not linked to a college admin profile.")
        logout(request)
        return redirect("college_auth")

    try:
        college = college_admin.college
        professors = Professor.objects.filter(college=college)
        labs = Lab.objects.filter(college=college)
    except DatabaseError:
        messages.error(request, "Database is not ready. Run migrations and verify DATABASE_URL.")
        return redirect("college_auth")

    return render(
        request,
        "college/dashboard.html",
        {
            "professors": professors,
            "labs": labs
        }
    )

def college_logout(request):
    logout(request)
    return redirect("college_auth")

def professor_auth(request):
    # determine which tab should be active; default to login
    active_tab = "login"

    if request.method == "POST":
        # LOGIN
        if "login_submit" in request.POST:
            username = request.POST.get("username")
            password = request.POST.get("password")

            user = authenticate(request, username=username, password=password)

            if user is not None and hasattr(user, "professor"):
                login(request, user)
                return redirect("professor_dashboard")

            else:
                messages.error(request, "Invalid username or password")

        # REGISTER
        elif "register_submit" in request.POST:
            active_tab = "register"

            username = request.POST.get("username")
            email = request.POST.get("email")
            password1 = request.POST.get("password1")
            password2 = request.POST.get("password2")
            divisions = request.POST.getlist("divisions")

            if password1 != password2:
                messages.error(request, "Passwords do not match")
                return render(request, "professor/professor_login.html", {"active_tab": active_tab})

            # Check if email exists in Professor model
            try:
                professor = Professor.objects.get(email=email)
            except Professor.DoesNotExist:
                messages.error(request, "Email not found in professor records. Contact your college admin.")
                return render(request, "professor/professor_login.html", {"active_tab": active_tab})

            # Check if this professor profile is already linked to a User
            if professor.user:
                messages.error(request, "An account is already created for this email.")
                return render(request, "professor/professor_login.html", {"active_tab": active_tab})

            # Check if the desired username is already taken by any User
            if User.objects.filter(username=username).exists():
                messages.error(request, "This username is already taken. Please choose another one.")
                return render(request, "professor/professor_login.html", {"active_tab": active_tab})

            user = User.objects.create_user(
                username=username,
                password=password1
            )

            professor.user = user
            professor.divisions = ','.join(divisions) if divisions else ''
            professor.save()

            messages.success(request, "Account created successfully. Please login.")
            active_tab = "login"

    return render(request, "professor/professor_login.html", {"active_tab": active_tab})

@login_required
def view_professors(request):

    college = request.user.collegeadmin.college

    professors = Professor.objects.filter(college=college)

    return render(
        request,
        "college/view_professors.html",
        {
            "professors": professors
        }
    )


    
@login_required
def add_professor(request):
    if request.method != "POST":
        return redirect("college_dashboard")

    try:
        college = request.user.collegeadmin.college
    except ObjectDoesNotExist:
        messages.error(request, "Your account is not linked to a college admin profile.")
        logout(request)
        return redirect("college_auth")

    name = request.POST.get("name", "").strip()
    email = request.POST.get("email", "").strip()
    course = request.POST.get("course", "").strip()

    if not name or not email or not course:
        messages.error(request, "Name, email, and course are required.")
        return redirect("college_dashboard")

    if Professor.objects.filter(email__iexact=email).exists():
        messages.error(request, "A professor with this email already exists.")
        return redirect("college_dashboard")

    try:
        Professor.objects.create(
            college=college,
            name=name,
            email=email,
            course=course
        )
    except IntegrityError:
        messages.error(request, "Could not add professor due to duplicate data.")
        return redirect("college_dashboard")
    except DatabaseError:
        messages.error(request, "Database is not ready. Run migrations and verify DATABASE_URL.")
        return redirect("college_auth")

    messages.success(request, f"Professor {name} added successfully!")
    return redirect("college_dashboard")



@login_required
def edit_profile(request):

    if request.method == "POST":

        user = request.user

        current_password = request.POST.get("currentPassword")
        username = request.POST.get("newUsername")
        new_password = request.POST.get("newPassword")
        confirm_password = request.POST.get("confirmPassword")

        # Check current password
        if not user.check_password(current_password):
            messages.error(request, "Current password is incorrect.")
            return redirect("college_dashboard")

        if username:
            user.username = username
            user.save()

        if new_password:
            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Password updated successfully.")
            else:
                messages.error(request, "New passwords do not match.")
                return redirect("college_dashboard")

        messages.success(request, "Profile updated successfully.")
        return redirect("college_dashboard")
    
@login_required
def delete_professor(request, professor_id):
    try:
        professor = Professor.objects.get(id=professor_id)
        
        # 1. Delete associated ExcelUpload files
        uploads = ExcelUpload.objects.filter(professor=professor)
        for upload in uploads:
            if upload.file:
                upload.file.delete()
        
        # 2. Handle Labs: Clear student mappings, delete manual/syllabus files
        labs = Lab.objects.filter(professor=professor)
        for lab in labs:
            lab.students.clear()  # Drop the many-to-many relationship
            if lab.syllabus:
                lab.syllabus.delete()
            if lab.manual:
                lab.manual.delete()
                
        # 3. Store the linked user before deleting the professor
        linked_user = professor.user
        
        # 4. Delete the professor (this cascaded to Lab, ExcelUpload DB records)
        professor.delete()
        
        # 5. Delete the User auth record so the username is freed up
        if linked_user:
            linked_user.delete()
            
    except Professor.DoesNotExist:
        pass

    return redirect("college_dashboard")

@login_required
def edit_professor(request):

    if request.method == "POST":

        professor_id = request.POST.get("professor_id")

        name = request.POST.get("name")

        email = request.POST.get("email")

        course = request.POST.get("course")

        professor = Professor.objects.get(id=professor_id)

        professor.name = name
        professor.email = email
        professor.course = course

        professor.save()

    return redirect("college_dashboard")

@login_required
def add_lab(request):

    if request.method == "POST":

        lab_name = request.POST.get("lab_name")

        professor_id = request.POST.get("professor")

        college = request.user.collegeadmin.college

        professor = Professor.objects.get(id=professor_id)

        Lab.objects.create(

            name=lab_name,

            professor=professor,

            college=college

        )

        messages.success(request, "Lab created successfully")

        return redirect("college_dashboard")
    
def professor_register(request):

    if request.method == "POST" and "register_submit" in request.POST:
        
        email = request.POST.get("email")
        username = request.POST.get("username")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        divisions = request.POST.getlist("divisions")

        if password1 != password2:
            messages.error(request, "Passwords do not match")
            return redirect("professor_login")

        try:
            professor = Professor.objects.get(email=email)

        except Professor.DoesNotExist:
            messages.error(request, "This email is not assigned by admin")
            return redirect("professor_auth")

        if professor.user:
            messages.error(request, "Account already created for this email")
            return redirect("professor_auth")
            
        if User.objects.filter(username=username).exists():
            messages.error(request, "This username is already taken. Please choose another one.")
            return redirect("professor_auth")

        user = User.objects.create_user(
            username=username,
            password=password1,
            email=email
        )

        professor.user = user
        professor.divisions = ",".join(divisions)
        professor.save()

        messages.success(request, "Account created successfully")

        return redirect("professor_login")

    return redirect("professor_login")   

def professor_login(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:

            try:
                Professor.objects.get(user=user)

                login(request, user)

                return redirect("professor_dashboard")

            except Professor.DoesNotExist:

                messages.error(request, "Not a professor account")

        else:
            messages.error(request, "Invalid login")

    return render(request, "professor/professor_login.html")

@login_required
def professor_dashboard(request):

    professor = Professor.objects.get(user=request.user)
    
    # Parse divisions from comma-separated string
    divisions_list = [d.strip() for d in professor.divisions.split(',') if d.strip()] if professor.divisions else []

    # Parse courses from the professor's profile
    course_names = [c.strip() for c in professor.course.split(',') if c.strip()] if professor.course else []
    
    # Ensure Lab objects exist for each assigned course
    assigned_labs = []
    for name in course_names:
        lab, created = Lab.objects.get_or_create(
            professor=professor,
            college=professor.college,
            name=name
        )
        assigned_labs.append(lab)

    # Get all divisions for this college
    all_divisions = Division.objects.filter(college=professor.college).order_by('name')

    # Fetch all pending viva submissions for the professor's labs
    pending_vivas = Submission.objects.filter(
        lab__in=assigned_labs,
        status='pending'
    ).select_related('student__user', 'experiment').order_by('-submitted_at')

    return render(
        request,
        "professor/professor_dashboard.html",
        {
            "professor": professor,
            "professor_divisions": divisions_list,
            "assigned_labs": assigned_labs,
            "all_divisions": all_divisions,
            "pending_vivas": pending_vivas
        }
    )

@login_required
def professor_logout(request):
    logout(request)
    return redirect("professor_auth")

@login_required
def professor_edit_profile(request):
    if request.method == "POST":
        user = request.user
        professor = Professor.objects.get(user=user)
        
        # Handle profile picture upload
        if 'profile_pic' in request.FILES:
            profile_pic = request.FILES['profile_pic']
            # Save the image to media folder (requires MEDIA_URL and MEDIA_ROOT configured)
            # For now, we'll just show a success message
            messages.success(request, "Profile updated successfully!")
        else:
            messages.success(request, "Profile updated successfully!")
        
        return redirect("professor_dashboard")
    
    return redirect("professor_dashboard")

# New views for enhanced functionality

@csrf_exempt
@require_POST
@login_required
def upload_student_excel(request):
    """Upload student excel file for a division and assign to a specific lab"""
    try:
        professor = Professor.objects.get(user=request.user)
        excel_file = request.FILES.get('excel_file')
        division_name = request.POST.get('division', '').strip()
        lab_id = request.POST.get('lab_id')

        if not excel_file or not division_name or not lab_id:
            return JsonResponse({'success': False, 'error': 'Missing file, division, or lab selection'}, status=400)

        if not excel_file.name.lower().endswith('.xlsx'):
            return JsonResponse({'success': False, 'error': 'Please upload a valid .xlsx file.'}, status=400)

        # Get the lab
        try:
            lab = Lab.objects.get(id=lab_id, professor=professor)
        except Lab.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Invalid lab selected'}, status=400)

        # Get or create division
        division, created = Division.objects.get_or_create(
            name=division_name,
            college=professor.college
        )

        # Check if an upload already exists for this Lab and Division
        if ExcelUpload.objects.filter(professor=professor, lab=lab, division=division).exists():
            return JsonResponse({
                'success': False, 
                'error': f'An Excel file is already uploaded for {lab.name} - {division_name}. Please delete the existing upload first.'
            }, status=409)

        # Guard file size on memory-constrained instances (Render starter/free).
        max_excel_size_bytes = 3 * 1024 * 1024  # 3 MB
        if getattr(excel_file, 'size', 0) > max_excel_size_bytes:
            return JsonResponse(
                {
                    'success': False,
                    'error': 'Excel file is too large for current server plan. Keep it under 3 MB.'
                },
                status=400
            )

        # Load workbook in read-only mode to reduce memory usage.
        wb = None
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        except Exception:
            return JsonResponse({'success': False, 'error': 'Could not read this Excel file. Use a valid .xlsx file.'}, status=400)

        parsed_rows = []
        seen_prns = set()
        try:
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if not row or len(row) < 2 or not row[0] or not row[1]:
                    continue

                name = str(row[0]).strip()
                prn = str(row[1]).strip()
                if not prn or prn.lower() == 'none' or prn in seen_prns:
                    continue

                seen_prns.add(prn)
                parsed_rows.append((name, prn))
        finally:
            if wb is not None:
                wb.close()

        if not parsed_rows:
            return JsonResponse({'success': False, 'error': 'No valid student rows found in the Excel file.'}, status=400)

        total_processed = 0
        conflicts_skipped = 0

        with transaction.atomic():
            prns = [prn for _, prn in parsed_rows]
            usernames = [prn.lower() for prn in prns]
            name_by_username = {prn.lower(): name for name, prn in parsed_rows}

            existing_users = {u.username: u for u in User.objects.filter(username__in=usernames)}
            missing_usernames = [uname for uname in usernames if uname not in existing_users]

            if missing_usernames:
                new_users = []
                for username in missing_usernames:
                    full_name = name_by_username.get(username, '').strip()
                    parts = full_name.split()
                    first_name = parts[0] if parts else ''
                    last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
                    user = User(username=username, first_name=first_name, last_name=last_name)
                    user.set_unusable_password()
                    new_users.append(user)

                User.objects.bulk_create(new_users, ignore_conflicts=True)
                existing_users = {u.username: u for u in User.objects.filter(username__in=usernames)}

            existing_students = {s.prn: s for s in Student.objects.filter(prn__in=prns)}
            linked_students = {
                s.user_id: s for s in Student.objects.filter(user_id__in=[u.id for u in existing_users.values() if u.id])
            }

            students_to_create = []
            students_to_update = []

            for _, prn in parsed_rows:
                username = prn.lower()
                user = existing_users.get(username)
                if not user:
                    continue

                student = existing_students.get(prn)
                if student:
                    if student.division_id != division.id:
                        student.division = division
                        students_to_update.append(student)
                    continue

                # Avoid one-to-one conflicts: same user linked to another PRN.
                linked_student = linked_students.get(user.id)
                if linked_student and linked_student.prn != prn:
                    conflicts_skipped += 1
                    continue

                students_to_create.append(Student(prn=prn, user=user, division=division))

            if students_to_create:
                Student.objects.bulk_create(students_to_create, ignore_conflicts=True)

            if students_to_update:
                Student.objects.bulk_update(students_to_update, ['division'])

            assigned_students = list(Student.objects.filter(prn__in=prns))
            through_model = lab.students.through
            through_rows = [through_model(lab_id=lab.id, student_id=stu.id) for stu in assigned_students]
            if through_rows:
                through_model.objects.bulk_create(through_rows, ignore_conflicts=True)

            total_processed = len(assigned_students)

        # Update professor's divisions list if new
        division_added = False
        current_divisions = [d.strip() for d in professor.divisions.split(',') if d.strip()]
        if division_name not in current_divisions:
            current_divisions.append(division_name)
            professor.divisions = ','.join(current_divisions)
            professor.save()
            division_added = True

        # Save upload tracking. If file storage fails, keep processing outcome
        # and save filename-only marker to avoid blocking student imports.
        try:
            excel_file.seek(0)
            ExcelUpload.objects.create(
                professor=professor,
                lab=lab,
                division=division,
                file=excel_file,
                filename=excel_file.name
            )
        except Exception:
            logger.exception(
                "ExcelUpload file save failed for professor_id=%s, lab_id=%s, division=%s",
                professor.id,
                lab.id,
                division.name,
            )
            ExcelUpload.objects.create(
                professor=professor,
                lab=lab,
                division=division,
                file=f"excel_uploads/{excel_file.name}",
                filename=excel_file.name
            )

        message = f'Successfully processed {total_processed} students and assigned to division {division_name}.'
        if conflicts_skipped:
            message += f' Skipped {conflicts_skipped} conflicted rows (same user linked to another PRN).'

        return JsonResponse({
            'success': True,
            'students_count': total_processed,
            'division_added': division_added,
            'message': message
        })

    except Professor.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Professor profile not found for this account.'}, status=403)
    except Exception as e:
        logger.exception("Student Excel upload failed")
        return JsonResponse(
            {
                'success': False,
                'error': f'{e.__class__.__name__}: {str(e)}'
            },
            status=500
        )

@login_required
def get_students_for_division(request):
    """Get students for a division and lab for marks entry"""
    division_name = request.GET.get('division', '').strip()
    lab_id = request.GET.get('lab_id')
    
    if not division_name or not lab_id:
        return JsonResponse({'students': []})

    try:
        professor = Professor.objects.get(user=request.user)
        # Use __iexact for safer division name matching
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id)
        
        # Filter students who are in this division AND assigned to this lab
        students = Student.objects.filter(division=division, assigned_labs=lab).order_by('prn')
        print(f"DEBUG: Found {students.count()} students for this division/lab")

        date = request.GET.get('date')
        
        students_data = []
        for student in students:
            # Get existing evaluations for this student in this lab
            marks = {}
            from django.db.models import Q
            submissions = Submission.objects.filter(
                Q(experiment__lab=lab) | Q(lab=lab), 
                student=student
            ).select_related('experiment', 'evaluation')
            
            for sub in submissions:
                eval_obj = getattr(sub, 'evaluation', None)
                if eval_obj:
                    exp_num = None
                    if sub.experiment:
                        exp_num = sub.experiment.number
                    elif sub.experiment_name:
                        import re
                        match = re.search(r'\d+', sub.experiment_name)
                        if match:
                            exp_num = int(match.group())
                    
                    if exp_num is not None:
                        marks[str(exp_num)] = {
                            'viva': float(eval_obj.viva_marks),
                            'marks': float(eval_obj.experiment_marks),
                            'writing': float(eval_obj.writeup_marks)
                        }
            
            student_dict = {
                'id': student.id,
                'name': student.user.get_full_name() or student.user.username,
                'prn': student.prn,
                'marks': marks,
                'attendance': None # Default is not marked
            }

            # Get attendance for specific date if provided
            if date:
                try:
                    att = Attendance.objects.filter(student=student, lab=lab, date=date).first()
                    if att:
                        student_dict['attendance'] = att.present
                except Exception:
                    pass

            students_data.append(student_dict)

        return JsonResponse({'students': students_data})
    except Exception as e:
        return JsonResponse({'students': [], 'error': str(e)})

@login_required
def get_submissions_for_division(request):
    """Get submissions for a division and lab for view uploads"""
    division_name = request.GET.get('division', '').strip()
    lab_id = request.GET.get('lab_id')
    if not division_name or not lab_id:
        return JsonResponse({'submissions': []})

    try:
        professor = Professor.objects.get(user=request.user)
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id)
        
        # Fetch both new-style (lab FK) and old-style (experiment FK) submissions.
        students = Student.objects.filter(division=division, assigned_labs=lab)
        from django.db.models import Q
        submissions = Submission.objects.filter(
            student__in=students
        ).filter(
            Q(lab=lab) | Q(experiment__lab=lab)
        ).select_related('student', 'student__user', 'experiment').order_by('student__prn', 'submitted_at')

        submissions_data = []
        for submission in submissions:
            code_field = submission.code_screenshot
            output_field = submission.output_screenshot

            code_name = getattr(code_field, 'name', '') if code_field else ''
            output_name = getattr(output_field, 'name', '') if output_field else ''

            # Skip marks-only rows that do not include any uploaded files.
            if not (code_name or output_name):
                continue

            code_url = _safe_field_url(code_field)
            output_url = _safe_field_url(output_field)

            if submission.experiment_name:
                experiment_label = submission.experiment_name
            elif submission.experiment:
                exp_title = submission.experiment.title or ''
                if exp_title:
                    experiment_label = f"Experiment {submission.experiment.number} - {exp_title}"
                else:
                    experiment_label = f"Experiment {submission.experiment.number}"
            else:
                experiment_label = ''

            submissions_data.append({
                'id': submission.id,
                'student_name': submission.student.user.get_full_name() or submission.student.user.username,
                'student_prn': submission.student.prn,
                'experiment_name': experiment_label,
                'experiment_title': experiment_label,
                'code_screenshot': code_url,
                'output_screenshot': output_url,
                'submitted_at': submission.submitted_at.strftime('%d %b %Y %I:%M %p'),
                'status': submission.status
            })

        return JsonResponse({'submissions': submissions_data})
    except Exception as e:
        return JsonResponse({'submissions': [], 'error': str(e)})

@login_required
def check_upload_status(request):
    """Check if an Excel file has already been uploaded for a lab/division"""
    lab_id = request.GET.get('lab_id')
    division_name = request.GET.get('division', '').strip()
    
    if not lab_id or not division_name:
        return JsonResponse({'exists': False})

    try:
        professor = Professor.objects.get(user=request.user)
        # We search by name and college to find the right division
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        
        upload = ExcelUpload.objects.filter(
            professor=professor,
            lab_id=lab_id,
            division=division
        ).first()

        if upload:
            return JsonResponse({
                'exists': True,
                'filename': upload.filename,
                'uploaded_at': timezone.localtime(upload.uploaded_at).strftime('%d %b %Y %I:%M %p')
            })
    except Exception:
        pass

    return JsonResponse({'exists': False})

@csrf_exempt
@require_POST
@login_required
def delete_upload(request):
    """Delete an existing Excel upload and reset the student assignments for that lab"""
    try:
        data = json.loads(request.body)
        lab_id = data.get('lab_id')
        division_name = data.get('division', '').strip()
        
        if not lab_id or not division_name:
            return JsonResponse({'success': False, 'error': 'Missing lab or division information'})

        professor = Professor.objects.get(user=request.user)
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id, professor=professor)

        upload = ExcelUpload.objects.filter(
            professor=professor,
            lab=lab,
            division=division
        ).first()

        if upload:
            # 1. Find students from this division
            students_in_division = Student.objects.filter(division=division)
            
            # 2. Delete all their submissions for THIS specific lab
            # This will automatically delete Evaluation records due to OneToOne CASCADE
            Submission.objects.filter(student__in=students_in_division, experiment__lab=lab).delete()
            
            # 3. Remove these students from the lab's student list (reset assignments)
            for student in students_in_division:
                lab.students.remove(student)
            
            # 4. Delete the actual file from storage
            if upload.file:
                upload.file.delete()
            
            # 5. Delete the tracking record
            upload.delete()

            return JsonResponse({'success': True, 'message': 'Upload deleted, marks cleared, and lab assignments reset.'})
        
        return JsonResponse({'success': False, 'error': 'No upload record found to delete.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_POST
@login_required(login_url='professor_auth')
def save_marks(request):
    """Save marks for students using batched queries to avoid timeouts."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    division_name = (data.get('division') or '').strip()
    lab_id = data.get('lab_id')
    marks_data = data.get('marks_data')

    if not division_name or lab_id is None or not isinstance(marks_data, list):
        return JsonResponse(
            {'success': False, 'error': 'Missing or invalid division, lab, or marks data'},
            status=400
        )

    def _to_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    try:
        professor = Professor.objects.get(user=request.user)
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id, professor=professor, college=professor.college)
    except Professor.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Professor profile not found.'}, status=403)
    except Division.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Division not found for this college.'}, status=404)
    except Lab.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Lab not found for this professor.'}, status=404)

    prns = []
    experiment_numbers = set()
    for item in marks_data:
        if not isinstance(item, dict):
            continue
        prn = str(item.get('prn', '')).strip()
        if prn:
            prns.append(prn)

        experiments = item.get('experiments', [])
        if not isinstance(experiments, list):
            continue

        for exp_data in experiments:
            if not isinstance(exp_data, dict):
                continue
            try:
                exp_num = int(exp_data.get('experiment_number'))
            except (TypeError, ValueError):
                continue
            if exp_num > 0:
                experiment_numbers.add(exp_num)

    if not prns or not experiment_numbers:
        return JsonResponse({'success': True, 'saved_count': 0, 'updated_submissions': 0})

    students = Student.objects.filter(
        prn__in=prns,
        division=division,
        assigned_labs=lab
    ).only('id', 'prn')
    students_by_prn = {student.prn: student for student in students}

    if not students_by_prn:
        return JsonResponse(
            {'success': False, 'error': 'No valid students found for this division and lab.'},
            status=400
        )

    saved_count = 0
    updated_submissions = 0

    try:
        with transaction.atomic():
            existing_experiments = Experiment.objects.filter(
                lab=lab,
                number__in=experiment_numbers
            )
            experiments_by_number = {
                experiment.number: experiment for experiment in existing_experiments
            }

            missing_experiments = [
                Experiment(lab=lab, number=exp_num, title=f"Experiment {exp_num}")
                for exp_num in experiment_numbers
                if exp_num not in experiments_by_number
            ]
            if missing_experiments:
                Experiment.objects.bulk_create(missing_experiments)
                for experiment in Experiment.objects.filter(lab=lab, number__in=experiment_numbers):
                    experiments_by_number[experiment.number] = experiment

            student_ids = [student.id for student in students_by_prn.values()]
            experiment_ids = [experiment.id for experiment in experiments_by_number.values()]

            existing_submissions = Submission.objects.filter(
                student_id__in=student_ids,
                experiment_id__in=experiment_ids
            ).only('id', 'student_id', 'experiment_id', 'status')
            submissions_by_key = {
                (submission.student_id, submission.experiment_id): submission
                for submission in existing_submissions
            }

            work_items = []
            missing_submissions = []

            for item in marks_data:
                if not isinstance(item, dict):
                    continue

                prn = str(item.get('prn', '')).strip()
                student = students_by_prn.get(prn)
                if student is None:
                    continue

                experiments = item.get('experiments', [])
                if not isinstance(experiments, list):
                    continue

                for exp_data in experiments:
                    if not isinstance(exp_data, dict):
                        continue

                    try:
                        exp_num = int(exp_data.get('experiment_number'))
                    except (TypeError, ValueError):
                        continue

                    experiment = experiments_by_number.get(exp_num)
                    if experiment is None:
                        continue

                    viva = _to_float(exp_data.get('viva_marks', 0))
                    exp_marks = _to_float(exp_data.get('experiment_marks', 0))
                    writeup = _to_float(exp_data.get('writeup_marks', 0))

                    key = (student.id, experiment.id)
                    submission = submissions_by_key.get(key)

                    # Do not create empty mark records for non-submitted work.
                    if submission is None and viva == 0 and exp_marks == 0 and writeup == 0:
                        continue

                    if submission is None:
                        missing_submissions.append(
                            Submission(
                                student=student,
                                lab=lab,
                                experiment=experiment,
                                experiment_name=f"Experiment {exp_num}",
                                code_screenshot='',
                                output_screenshot='',
                                status='evaluated'
                            )
                        )

                    work_items.append((key, viva, exp_marks, writeup))

            if missing_submissions:
                Submission.objects.bulk_create(missing_submissions)
                # Reload to ensure IDs are available on all DB backends.
                submissions_by_key = {
                    (submission.student_id, submission.experiment_id): submission
                    for submission in Submission.objects.filter(
                        student_id__in=student_ids,
                        experiment_id__in=experiment_ids
                    ).only('id', 'student_id', 'experiment_id', 'status')
                }

            target_submission_ids = list({
                submissions_by_key[key].id
                for key, _, _, _ in work_items
                if key in submissions_by_key
            })

            evaluations_by_submission_id = {
                evaluation.submission_id: evaluation
                for evaluation in Evaluation.objects.filter(
                    submission_id__in=target_submission_ids
                )
            }

            evaluations_to_create = []
            evaluations_to_update = []
            submissions_to_update = {}

            for key, viva, exp_marks, writeup in work_items:
                submission = submissions_by_key.get(key)
                if submission is None:
                    continue

                evaluation = evaluations_by_submission_id.get(submission.id)
                if evaluation is None:
                    evaluations_to_create.append(
                        Evaluation(
                            submission=submission,
                            viva_marks=viva,
                            experiment_marks=exp_marks,
                            writeup_marks=writeup
                        )
                    )
                else:
                    changed = False
                    if evaluation.viva_marks != viva:
                        evaluation.viva_marks = viva
                        changed = True
                    if evaluation.experiment_marks != exp_marks:
                        evaluation.experiment_marks = exp_marks
                        changed = True
                    if evaluation.writeup_marks != writeup:
                        evaluation.writeup_marks = writeup
                        changed = True
                    if changed:
                        evaluations_to_update.append(evaluation)

                if submission.status != 'evaluated':
                    submission.status = 'evaluated'
                    submissions_to_update[submission.id] = submission

                saved_count += 1

            if evaluations_to_create:
                Evaluation.objects.bulk_create(evaluations_to_create)
            if evaluations_to_update:
                Evaluation.objects.bulk_update(
                    evaluations_to_update,
                    ['viva_marks', 'experiment_marks', 'writeup_marks']
                )
            if submissions_to_update:
                Submission.objects.bulk_update(
                    list(submissions_to_update.values()),
                    ['status']
                )
                updated_submissions = len(submissions_to_update)

        return JsonResponse({
            'success': True,
            'saved_count': saved_count,
            'updated_submissions': updated_submissions
        })
    except Exception as e:
        logger.exception("save_marks failed")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def download_marks_excel(request):
    """Download marks as Excel file"""
    division_name = (request.GET.get('division') or '').strip()
    lab_id = request.GET.get('lab_id')
    if not division_name or not lab_id:
        return HttpResponse("Division or Lab not specified", status=400)

    try:
        professor = Professor.objects.get(user=request.user)
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id, professor=professor, college=professor.college)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _safe_worksheet_title(f"Marks_{division.name}", fallback="Marks")
        
        # Headers
        headers = ['PRN', 'Student Name']
        for exp in range(1, 11): # Assuming up to 10 experiments per lab
            headers.extend([f'Exp{exp}_Viva', f'Exp{exp}_Marks', f'Exp{exp}_Writing', f'Exp{exp}_Total'])
        headers.extend(['Total_Marks', 'Average_Marks'])
        ws.append(headers)

        # Fetch students once and preload all relevant submissions/evaluations in one query.
        students = list(
            Student.objects.filter(division=division, assigned_labs=lab)
            .select_related('user')
            .order_by('prn')
        )
        student_ids = [student.id for student in students]

        marks_by_student_exp = {}
        if student_ids:
            submissions = (
                Submission.objects.filter(student_id__in=student_ids, experiment__lab=lab)
                .select_related('experiment', 'evaluation')
                .order_by('student_id', 'experiment__number', 'id')
            )
            for submission in submissions.iterator(chunk_size=2000):
                experiment = getattr(submission, 'experiment', None)
                if not experiment:
                    continue
                exp_num = int(experiment.number or 0)
                if exp_num < 1 or exp_num > 10:
                    continue

                eval_obj = getattr(submission, 'evaluation', None)
                if not eval_obj:
                    continue

                key = (submission.student_id, exp_num)
                # Keep first submission per student/experiment to match prior behavior.
                if key in marks_by_student_exp:
                    continue

                viva = float(eval_obj.viva_marks or 0)
                exp_marks = float(eval_obj.experiment_marks or 0)
                writing = float(eval_obj.writeup_marks or 0)
                marks_by_student_exp[key] = (viva, exp_marks, writing, viva + exp_marks + writing)

        for student in students:
            row = [student.prn, student.user.get_full_name() or student.user.username]
            
            total_marks = float(0)
            exp_count = float(0)
            
            for exp_num in range(1, 11):
                data = marks_by_student_exp.get((student.id, exp_num))
                if data:
                    viva, exp_marks, writing, exp_total = data
                    row.extend([viva, exp_marks, writing, exp_total])
                    total_marks = total_marks + exp_total
                    exp_count = exp_count + 1
                else:
                    row.extend([0, 0, 0, 0])
            
            average = float(total_marks) / float(exp_count) if exp_count > 0 else 0.0
            row.extend([total_marks, round(average, 2)])
            ws.append(row)
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        division_slug = _safe_filename_part(division.name, fallback='division')
        lab_slug = _safe_filename_part(lab.name, fallback='lab')
        response['Content-Disposition'] = f'attachment; filename="marks_{division_slug}_{lab_slug}.xlsx"'
        return response
    except Professor.DoesNotExist:
        return HttpResponse("Professor profile not found.", status=403)
    except Division.DoesNotExist:
        return HttpResponse("Division not found for this college.", status=404)
    except Lab.DoesNotExist:
        return HttpResponse("Lab not found for this professor.", status=404)
    except Exception as e:
        logger.exception("download_marks_excel failed")
        return HttpResponse(f"Export failed: {str(e)}", status=500)

@login_required
def download_total_marks_excel(request):
    """Download Excel with only total marks per experiment per student (no breakdown)"""
    division_name = (request.GET.get('division') or '').strip()
    lab_id = request.GET.get('lab_id')

    if not division_name or not lab_id:
        return HttpResponse("Missing division or lab_id", status=400)

    try:
        professor = Professor.objects.get(user=request.user)
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id, professor=professor, college=professor.college)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _safe_worksheet_title(f"Totals_{division.name}", fallback="Totals")

        from openpyxl.styles import Font, Alignment, PatternFill
        bold_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_font = Font(bold=True, color="FFFFFF")

        # Build header: Student Name | PRN | Exp 1 Total | ... | Exp 10 Total | Grand Total
        headers = ['Student Name', 'PRN']
        for exp in range(1, 11):
            headers.append(f'Exp {exp} Total (15)')
        headers.append('Grand Total')

        # Write header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        students = list(
            Student.objects.filter(division=division, assigned_labs=lab)
            .select_related('user')
            .order_by('prn')
        )
        student_ids = [student.id for student in students]

        totals_by_student_exp = {}
        if student_ids:
            submissions = (
                Submission.objects.filter(student_id__in=student_ids, experiment__lab=lab)
                .select_related('experiment', 'evaluation')
                .order_by('student_id', 'experiment__number', 'id')
            )
            for submission in submissions.iterator(chunk_size=2000):
                experiment = getattr(submission, 'experiment', None)
                if not experiment:
                    continue
                exp_num = int(experiment.number or 0)
                if exp_num < 1 or exp_num > 10:
                    continue

                eval_obj = getattr(submission, 'evaluation', None)
                if not eval_obj:
                    continue

                key = (submission.student_id, exp_num)
                # Keep first submission per student/experiment to match prior behavior.
                if key in totals_by_student_exp:
                    continue

                exp_total = float(
                    (eval_obj.viva_marks or 0) +
                    (eval_obj.experiment_marks or 0) +
                    (eval_obj.writeup_marks or 0)
                )
                totals_by_student_exp[key] = exp_total

        for row_idx, student in enumerate(students, 2):
            row = [
                student.user.get_full_name() or student.user.username,
                student.prn
            ]
            grand_total = float(0)

            for exp_num in range(1, 11):
                exp_total = totals_by_student_exp.get((student.id, exp_num))
                if exp_total is not None:
                    grand_total = grand_total + exp_total
                    row.append(exp_total)
                else:
                    row.append(0)

            row.append(grand_total)
            ws.append(row)

            # Bold grand total cell
            ws.cell(row=row_idx, column=len(headers)).font = bold_font

        # Fixed widths avoid scanning entire columns for large classes.
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18
        for col_idx in range(3, len(headers)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        ws.column_dimensions[get_column_letter(len(headers))].width = 14

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        division_slug = _safe_filename_part(division.name, fallback='division')
        lab_slug = _safe_filename_part(lab.name, fallback='lab')
        response['Content-Disposition'] = f'attachment; filename="total_marks_{division_slug}_{lab_slug}.xlsx"'
        return response

    except Professor.DoesNotExist:
        return HttpResponse("Professor profile not found.", status=403)
    except Division.DoesNotExist:
        return HttpResponse("Division not found for this college.", status=404)
    except Lab.DoesNotExist:
        return HttpResponse("Lab not found for this professor.", status=404)
    except Exception as e:
        logger.exception("download_total_marks_excel failed")
        return HttpResponse(f"Export failed: {str(e)}", status=500)

@csrf_exempt
@require_POST
@login_required
def send_marks_report(request):
    """Send marks report as Excel attachment to professor's email"""
    try:
        data = json.loads(request.body)
        division_name = data.get('division')
        lab_id = data.get('lab_id')

        professor = Professor.objects.get(user=request.user)
        division = Division.objects.get(name=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id)
        students = Student.objects.filter(division=division).order_by('prn')

        from openpyxl.styles import Font, Alignment

        # --- Build Excel exactly like the individual marks export ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Marks - {division_name}"

        # Report title at the very top
        ws['A1'] = f"Marks Report — Division {division_name} | {lab.name}"
        ws['A1'].font = Font(bold=True, size=13)
        ws['A2'] = f"Professor: {professor.name}   |   Course: {professor.course}   |   Generated: {timezone.now().strftime('%d %b %Y %I:%M %p')}"
        ws['A2'].font = Font(italic=True)

        current_row = 4  # start below title

        col_headers = ["Exp No", "Experiment (5 M)", "Write-up (5 M)", "Viva (5 M)", "Total (15 M)"]

        for student in students:
            submissions = Submission.objects.filter(
                student=student, experiment__lab=lab
            ).select_related('evaluation', 'experiment').order_by('experiment__number')

            # Collect only evaluated submissions
            eval_subs = [s for s in submissions if hasattr(s, 'evaluation')]
            if not eval_subs:
                continue  # skip students with nothing evaluated

            # Student details row (same as individual export)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(row=current_row, column=1, value="Student Details").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=f"Name: {student.user.username}")
            ws.cell(row=current_row, column=4, value=f"PRN: {student.prn}")
            current_row += 1

            # Blank row then column headers
            for col_num, header in enumerate(col_headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Data rows — exactly same as individual export
            total_obtained = float(0)
            total_possible = float(0)
            for sub in eval_subs:
                viva      = float(sub.evaluation.viva_marks or 0)
                exp_marks = float(sub.evaluation.experiment_marks or 0)
                writeup   = float(sub.evaluation.writeup_marks or 0)
                total     = viva + exp_marks + writeup

                ws.cell(row=current_row, column=1, value=sub.experiment.number).alignment = Alignment(horizontal='center')
                ws.cell(row=current_row, column=2, value=exp_marks).alignment   = Alignment(horizontal='center')
                ws.cell(row=current_row, column=3, value=writeup).alignment     = Alignment(horizontal='center')
                ws.cell(row=current_row, column=4, value=viva).alignment        = Alignment(horizontal='center')
                ws.cell(row=current_row, column=5, value=total).alignment       = Alignment(horizontal='center')
                total_obtained += total
                total_possible += 15
                current_row += 1

            # Summary rows — same as individual export
            current_row += 1  # blank gap
            avg     = round(total_obtained / max(1, total_possible / 15), 2)
            percent = round((total_obtained / max(1, total_possible)) * 100, 2)

            ws.cell(row=current_row,   column=4, value="Grand Total:").font = Font(bold=True)
            ws.cell(row=current_row,   column=5, value=total_obtained)
            ws.cell(row=current_row+1, column=4, value="Average:").font    = Font(bold=True)
            ws.cell(row=current_row+1, column=5, value=avg)
            ws.cell(row=current_row+2, column=4, value="Percentage:").font = Font(bold=True)
            ws.cell(row=current_row+2, column=5, value=f"{percent}%")

            current_row += 4  # gap before next student

        # Save to in-memory buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_data = buffer.getvalue()
        filename = f"marks_report_{division_name}_{lab.name}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

        # --- Send email with Excel attachment ---
        email = EmailMessage(
            subject=f"Marks Report — Division {division_name} | {lab.name}",
            body=(
                f"Dear {professor.name},\n\n"
                f"Please find the marks report for Division {division_name} ({lab.name}) attached.\n\n"
                f"Generated on: {timezone.now().strftime('%d %b %Y %I:%M %p')}\n\n"
                f"Regards,\nLab Management System"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[professor.email],
        )
        email.attach(filename, excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.send(fail_silently=False)

        return JsonResponse({'success': True, 'message': f'Report sent to {professor.email}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

from django.views.decorators.clickjacking import xframe_options_exempt

@login_required
@xframe_options_exempt
def evaluate_submission(request, submission_id):
    """Evaluate a specific submission"""
    try:
        submission = Submission.objects.get(id=submission_id)
        professor = Professor.objects.get(user=request.user)
        
        # Check if professor has access to this submission
        # Use direct lab FK (new-style) or fall back to experiment's lab (old-style)
        sub_lab = submission.lab if submission.lab else (
            submission.experiment.lab if submission.experiment else None
        )
        if sub_lab is None or sub_lab.professor != professor:
            return redirect("professor_dashboard")
        
        evaluation, created = Evaluation.objects.get_or_create(submission=submission)
        
        if request.method == "POST":
            viva_marks_raw = request.POST.get('viva_marks', 0)
            experiment_marks_raw = request.POST.get('experiment_marks', 0)
            writeup_marks_raw = request.POST.get('writeup_marks', 0)

            evaluation.viva_marks = float(viva_marks_raw) if viva_marks_raw else 0.0
            evaluation.experiment_marks = float(experiment_marks_raw) if experiment_marks_raw else 0.0
            evaluation.writeup_marks = float(writeup_marks_raw) if writeup_marks_raw else 0.0
            
            evaluation.comments = request.POST.get('comments', '')
            evaluation.save()
            
            submission.status = 'evaluated'
            submission.save()
            
            messages.success(request, "✅ Evaluation saved successfully!")
            # Redirect back to same page so the success alert is visible
            from django.urls import reverse
            return redirect(reverse('evaluate_submission', args=[submission_id]))
        
        code_url = ''
        output_url = ''
        if _resource_exists(submission.code_screenshot):
            code_url = _safe_field_url(submission.code_screenshot)
        if _resource_exists(submission.output_screenshot):
            output_url = _safe_field_url(submission.output_screenshot)

        return render(request, "professor/evaluate_submission.html", {
            'submission': submission,
            'evaluation': evaluation,
            'code_url': code_url,
            'output_url': output_url
        })
        
    except Submission.DoesNotExist:
        messages.error(request, "Submission not found")
        return redirect("professor_dashboard")

@login_required
def delete_submission(request, submission_id):
    """Allow students to delete their own pending submissions"""
    try:
        student = Student.objects.get(user=request.user)
        submission = Submission.objects.get(id=submission_id, student=student)
        
        if submission.status == 'pending':
            submission.delete()
            messages.success(request, "Submission deleted successfully.")
        else:
            messages.error(request, "Cannot delete an already evaluated submission.")
            
    except (Student.DoesNotExist, Submission.DoesNotExist):
        messages.error(request, "Submission not found.")
        
    return redirect('student_dashboard')

@login_required(login_url='professor_auth')
@csrf_exempt
def save_attendance(request):
    """Save/update attendance records for a division on a specific date"""
    if request.method != "POST":
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    try:
        data = json.loads(request.body)
        lab_id = data.get('lab_id')
        division_name = (data.get('division') or '').strip()
        date_str = data.get('date')
        attendance_data = data.get('attendance', [])  # [{student_id: int, present: bool}]

        if not all([lab_id, division_name, date_str]) or not isinstance(attendance_data, list):
            return JsonResponse({'success': False, 'error': 'Missing or invalid required fields'}, status=400)

        professor = Professor.objects.get(user=request.user)
        division = Division.objects.get(name__iexact=division_name, college=professor.college)
        lab = Lab.objects.get(id=lab_id, professor=professor, college=professor.college)
        # Parse date from YYYY-MM-DD
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()

        requested_student_ids = []
        normalized_entries = []
        for entry in attendance_data:
            if not isinstance(entry, dict):
                continue
            try:
                student_id = int(entry.get('student_id'))
            except (TypeError, ValueError):
                continue
            is_present = bool(entry.get('present'))
            requested_student_ids.append(student_id)
            normalized_entries.append((student_id, is_present))

        if not requested_student_ids:
            return JsonResponse({'success': False, 'error': 'No valid attendance rows provided.'}, status=400)

        valid_students = Student.objects.filter(
            id__in=requested_student_ids,
            division=division,
            assigned_labs=lab
        ).only('id')
        valid_student_ids = {student.id for student in valid_students}

        if not valid_student_ids:
            return JsonResponse({'success': False, 'error': 'No valid students found for this lab/division.'}, status=400)

        existing_records = Attendance.objects.filter(
            lab=lab,
            date=date_obj,
            student_id__in=valid_student_ids
        ).order_by('id')

        attendance_by_student_id = {}
        for record in existing_records:
            # Keep first record per student to tolerate historical duplicates.
            if record.student_id not in attendance_by_student_id:
                attendance_by_student_id[record.student_id] = record

        to_create = []
        to_update = []

        for student_id, is_present in normalized_entries:
            if student_id not in valid_student_ids:
                continue

            existing = attendance_by_student_id.get(student_id)
            if existing is None:
                to_create.append(
                    Attendance(
                        student_id=student_id,
                        lab=lab,
                        date=date_obj,
                        present=is_present
                    )
                )
            elif existing.present != is_present:
                existing.present = is_present
                to_update.append(existing)

        with transaction.atomic():
            if to_create:
                Attendance.objects.bulk_create(to_create)
            if to_update:
                Attendance.objects.bulk_update(to_update, ['present'])

        return JsonResponse({
            'success': True, 
            'message': f'Attendance for {date_str} saved successfully.',
            'saved_count': len(to_create) + len(to_update)
        })

    except Professor.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Professor profile not found.'}, status=403)
    except Division.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Division not found for this college.'}, status=404)
    except Lab.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Lab not found for this professor.'}, status=404)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)
    except Exception as e:
        logger.exception("save_attendance failed")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@csrf_exempt
@require_POST
def toggle_viva_session(request):
    """Activates or Deactivates a live Viva Session for a specific student"""
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        is_active = data.get('is_active')
        room_name = data.get('room_name', '')
        
        professor = Professor.objects.get(user=request.user)
        student = Student.objects.get(id=student_id)
        
        # Deactivate any existing active sessions to prevent ghost calls
        if is_active:
            VivaSession.objects.filter(professor=professor, student=student, is_active=True).update(is_active=False)
            
            # Create the new session
            VivaSession.objects.create(
                professor=professor,
                student=student,
                room_name=room_name,
                is_active=True
            )
            return JsonResponse({'success': True, 'message': 'Session activated'})
        else:
            # End the current session
            VivaSession.objects.filter(professor=professor, student=student, is_active=True).update(is_active=False)
            return JsonResponse({'success': True, 'message': 'Session ended'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='student_login')
def check_call(request):
    """Student dashboard pings this to see if a professor is calling them"""
    try:
        student = Student.objects.get(user=request.user)
        active_session = VivaSession.objects.filter(student=student, is_active=True).first()
        
        if active_session:
            return JsonResponse({
                'active': True,
                'room_name': active_session.room_name
            })
            
    except Student.DoesNotExist:
        pass
        
    return JsonResponse({'active': False})


@login_required
def submission_media(request, submission_id, media_type):
    """
    Serve submission screenshots securely for students and professors.
    """
    try:
        submission = Submission.objects.select_related('lab', 'experiment', 'student').get(id=submission_id)
    except Submission.DoesNotExist:
        return HttpResponse("Submission not found.", status=404)

    if not _submission_access_allowed(request, submission):
        return HttpResponse("You do not have access to this file.", status=403)

    if media_type == 'code':
        resource = submission.code_screenshot
    elif media_type == 'output':
        resource = submission.output_screenshot
    else:
        return HttpResponse("Invalid resource type.", status=400)

    if not _resource_exists(resource):
        return HttpResponse(
            "File is missing in storage for this submission.",
            status=404
        )

    return _file_response_from_field(resource, download=False)


@login_required
def download_lab_resource(request, lab_id, resource_type):
    """
    Shared endpoint for student/professor resource access.
    Students can access assigned labs; professors can access their own labs.
    """
    lab = Lab.objects.filter(id=lab_id).first()
    if lab is None:
        return HttpResponse("Lab not found.", status=404)

    is_assigned_student = Student.objects.filter(user=request.user, assigned_labs=lab).exists()
    is_lab_professor = Professor.objects.filter(user=request.user, id=lab.professor_id).exists()
    if not (is_assigned_student or is_lab_professor):
        return HttpResponse("You do not have access to this file.", status=403)

    if resource_type == 'syllabus':
        resource = lab.syllabus
    elif resource_type == 'manual':
        resource = lab.manual
    else:
        return HttpResponse("Invalid resource type.", status=400)

    if not _resource_exists(resource):
        return HttpResponse(
            "File is missing in storage for this lab resource. Please ask the professor to re-upload it.",
            status=404
        )

    should_download = request.GET.get('download') == '1'
    return _file_response_from_field(resource, download=should_download)

@login_required
def check_lab_resources_status(request):
    """Check if syllabus or manual are uploaded for a specific lab"""
    lab_id = request.GET.get('lab_id')
    if not lab_id:
        return JsonResponse({'success': False, 'error': 'Lab ID is required'})
    
    try:
        professor = Professor.objects.get(user=request.user)
        lab = Lab.objects.get(id=lab_id, professor=professor)
        
        syllabus_available = _resource_exists(lab.syllabus)
        manual_available = _resource_exists(lab.manual)

        return JsonResponse({
            'success': True,
            'syllabus_name': lab.syllabus.name.split('/')[-1] if syllabus_available else None,
            'syllabus_url': _build_lab_resource_url(lab.id, 'syllabus', download=False) if syllabus_available else None,
            'manual_name': lab.manual.name.split('/')[-1] if manual_available else None,
            'manual_url': _build_lab_resource_url(lab.id, 'manual', download=False) if manual_available else None
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_POST
@login_required
def upload_lab_resource(request):
    """Upload syllabus or lab manual for a specific lab"""
    lab_id = request.POST.get('lab_id')
    resource_type = request.POST.get('resource_type') # 'syllabus' or 'manual'
    uploaded_file = request.FILES.get('file')
    
    if not all([lab_id, resource_type, uploaded_file]):
        return JsonResponse({'success': False, 'error': 'Missing required fields'})
    
    try:
        professor = Professor.objects.get(user=request.user)
        lab = Lab.objects.get(id=lab_id, professor=professor)
        
        if resource_type == 'syllabus':
            # Delete old file if exists
            if lab.syllabus:
                lab.syllabus.delete()
            lab.syllabus = uploaded_file
        elif resource_type == 'manual':
            if lab.manual:
                lab.manual.delete()
            lab.manual = uploaded_file
        else:
            return JsonResponse({'success': False, 'error': 'Invalid resource type'})
            
        lab.save()
        return JsonResponse({
            'success': True, 
            'message': f'{resource_type.capitalize()} uploaded successfully',
            'filename': uploaded_file.name,
            'url': _build_lab_resource_url(lab.id, resource_type, download=False)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_POST
@login_required
def delete_lab_resource(request):
    """Delete syllabus or lab manual for a specific lab"""
    try:
        data = json.loads(request.body)
        lab_id = data.get('lab_id')
        resource_type = data.get('resource_type') # 'syllabus' or 'manual'
        
        if not all([lab_id, resource_type]):
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
            
        professor = Professor.objects.get(user=request.user)
        lab = Lab.objects.get(id=lab_id, professor=professor)
        
        if resource_type == 'syllabus':
            if lab.syllabus:
                lab.syllabus.delete()
                lab.syllabus = None
        elif resource_type == 'manual':
            if lab.manual:
                lab.manual.delete()
                lab.manual = None
        else:
            return JsonResponse({'success': False, 'error': 'Invalid resource type'})
            
        lab.save()
        return JsonResponse({'success': True, 'message': f'{resource_type.capitalize()} deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
